from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import cast, func, Integer, String
from typing import List, Optional
import asyncio
import json
from api.deps import get_db
from core.database import AsyncSessionLocal
from schemas.ipam import (
    IPSubnetCreate,
    IPSubnetUpdate,
    IPSubnetResponse,
    IPAddressResponse,
    IPAddressUpdate,
    IPAddressDetailResponse,
    IPAddressListResponse,
    IPScanRequest,
    IPScanSummary,
    IPAMScanStartResponse,
    IPAMScanStatusResponse,
    IPSubnetStatistics,
    IPAMDashboard,
    IPStatus,
    IPScanHistoryResponse,
    IPSubnetBatchImportRequest,
    IPSubnetBatchImportResult,
    NetworkSearchRequest,
    NetworkSearchResponse,
    SubnetCalculatorRequest,
    SubnetCalculatorResponse
)
from services.ipam_service import ipam_service
from services.ipam_scan_status import ipam_scan_status_service
from services.network_scheduler import network_scheduler
from models.ipam import IPAddress
from utils.logger import logger

# Temporarily disabled due to missing openpyxl package
# TODO: Install openpyxl when network is available
# import openpyxl
# from openpyxl import Workbook
# from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

router = APIRouter(prefix="/ipam", tags=["ipam"])


# Subnet endpoints
@router.post("/subnets", response_model=IPSubnetResponse, status_code=status.HTTP_201_CREATED)
async def create_subnet(
    subnet: IPSubnetCreate,
    db: AsyncSession = Depends(get_db)
):
    """
    Create a new IP subnet

    This will automatically generate all IP addresses in the subnet.
    """
    try:
        result = await ipam_service.create_subnet(
            db=db,
            name=subnet.name,
            network=subnet.network,
            description=subnet.description,
            vlan_id=subnet.vlan_id,
            gateway=str(subnet.gateway) if subnet.gateway else None,
            dns_servers=subnet.dns_servers,
            enabled=subnet.enabled,
            auto_scan=subnet.auto_scan,
            scan_interval=subnet.scan_interval
        )
        # Get statistics
        stats = await ipam_service.get_subnet_statistics(db, result.id)
        # Convert IP objects to strings
        return {
            "id": result.id,
            "name": result.name,
            "network": str(result.network),
            "description": result.description,
            "vlan_id": result.vlan_id,
            "gateway": str(result.gateway) if result.gateway else None,
            "dns_servers": result.dns_servers,
            "enabled": result.enabled,
            "auto_scan": result.auto_scan,
            "scan_interval": result.scan_interval,
            "last_scan_at": result.last_scan_at,
            "created_at": result.created_at,
            "updated_at": result.updated_at,
            **stats
        }
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        logger.error(f"Error creating subnet: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to create subnet: {str(e)}"
        )


@router.post("/subnets/batch", response_model=IPSubnetBatchImportResult, status_code=status.HTTP_201_CREATED)
async def batch_import_subnets(
    request: IPSubnetBatchImportRequest,
    db: AsyncSession = Depends(get_db)
):
    """
    Batch import multiple IP subnets

    This endpoint allows you to import multiple subnets at once.
    - Maximum 100 subnets per request
    - Optionally skip existing networks (duplicates)
    - Returns detailed results with success/failure counts
    """
    try:
        # Convert Pydantic models to dictionaries
        subnets_data = [subnet.model_dump() for subnet in request.subnets]

        # Call service method
        result = await ipam_service.batch_create_subnets(
            db=db,
            subnets_data=subnets_data,
            skip_existing=request.skip_existing
        )

        return result

    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        logger.error(f"Error in batch import: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Batch import failed: {str(e)}"
        )


@router.get("/subnets/template/download")
async def download_excel_template():
    """
    Download Excel template for batch subnet import

    Returns an Excel file with example data and column headers
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "子网导入模板"

        # Headers
        headers = ['子网名称', '网络地址(CIDR)', '描述', 'VLAN ID', '网关', 'DNS服务器']
        ws.append(headers)

        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Example data
        examples = [
            ['办公网络', '10.0.1.0/24', '办公区域网络', '100', '10.0.1.1', '8.8.8.8,8.8.4.4'],
            ['数据中心', '172.16.0.0/16', 'DC核心网络', '200', '172.16.0.1', '8.8.8.8'],
            ['访客网络', '192.168.10.0/24', '访客WiFi', '300', '192.168.10.1', ''],
        ]

        for row in examples:
            ws.append(row)

        # Adjust column widths
        column_widths = [20, 20, 25, 12, 15, 20]
        for idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + idx)].width = width

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=IPAM_Import_Template.xlsx"
            }
        )
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="Excel功能需要安装openpyxl库。请联系管理员执行: pip install openpyxl"
        )
    except Exception as e:
        logger.error(f"Error generating Excel template: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"生成Excel模板失败: {str(e)}"
        )


@router.post("/subnets/import/excel", response_model=IPSubnetBatchImportResult, status_code=status.HTTP_201_CREATED)
async def import_subnets_from_excel(
    file: UploadFile = File(...),
    skip_existing: bool = Query(True, description="Skip existing networks"),
    db: AsyncSession = Depends(get_db)
):
    """
    Import subnets from Excel file

    Upload an Excel file with subnet data. Use the template for correct format.
    Required columns: 子网名称, 网络地址(CIDR)
    Optional columns: 描述, VLAN ID, 网关, DNS服务器
    """
    try:
        from openpyxl import load_workbook
        from io import BytesIO

        # Validate file type
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="只支持 .xlsx 或 .xls 格式的Excel文件"
            )

        # Read file content
        contents = await file.read()
        workbook = load_workbook(BytesIO(contents), data_only=True)
        worksheet = workbook.active

        # Parse data (skip header row)
        subnets_data = []
        errors = []

        for idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            # Skip empty rows
            if not any(row):
                continue

            try:
                # Extract values (handle None values)
                name = row[0] if row[0] else f"子网-{idx}"
                network = row[1] if len(row) > 1 and row[1] else None
                description = row[2] if len(row) > 2 and row[2] else None
                vlan_id = int(row[3]) if len(row) > 3 and row[3] and str(row[3]).strip() else None
                gateway = row[4] if len(row) > 4 and row[4] else None
                dns_servers = row[5] if len(row) > 5 and row[5] else None

                # Validate required fields
                if not network:
                    errors.append({
                        'index': idx,
                        'network': str(name),
                        'error': '缺少网络地址(CIDR)'
                    })
                    continue

                subnet_data = {
                    'name': str(name).strip(),
                    'network': str(network).strip(),
                    'description': str(description).strip() if description else None,
                    'vlan_id': vlan_id,
                    'gateway': str(gateway).strip() if gateway else None,
                    'dns_servers': str(dns_servers).strip() if dns_servers else None,
                    'enabled': True,
                    'auto_scan': True,
                    'scan_interval': 3600
                }

                subnets_data.append(subnet_data)

            except Exception as e:
                errors.append({
                    'index': idx,
                    'network': str(row[1]) if len(row) > 1 else 'N/A',
                    'error': f'解析错误: {str(e)}'
                })

        if not subnets_data:
            return {
                'total': 0,
                'success': 0,
                'failed': len(errors),
                'skipped': 0,
                'errors': errors,
                'imported_ids': []
            }

        # Call batch import service
        result = await ipam_service.batch_create_subnets(
            db=db,
            subnets_data=subnets_data,
            skip_existing=skip_existing
        )

        # Merge parse errors with import errors
        result['errors'].extend(errors)
        result['failed'] += len(errors)

        return result

    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="Excel导入功能需要安装openpyxl库。请联系管理员执行: pip install openpyxl"
        )
    except Exception as e:
        logger.error(f"Error importing Excel: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Excel导入失败: {str(e)}"
        )


@router.get("/subnets", response_model=List[IPSubnetResponse])
async def list_subnets(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    db: AsyncSession = Depends(get_db)
):
    """List all IP subnets"""
    subnets = await ipam_service.list_subnets(db, skip=skip, limit=limit)

    # Add statistics to each subnet
    result = []
    for subnet in subnets:
        stats = await ipam_service.get_subnet_statistics(db, subnet.id)
        subnet_dict = {
            "id": subnet.id,
            "name": subnet.name,
            "network": str(subnet.network),
            "description": subnet.description,
            "vlan_id": subnet.vlan_id,
            "gateway": str(subnet.gateway) if subnet.gateway else None,
            "dns_servers": subnet.dns_servers,
            "enabled": subnet.enabled,
            "auto_scan": subnet.auto_scan,
            "scan_interval": subnet.scan_interval,
            "last_scan_at": subnet.last_scan_at,
            "created_at": subnet.created_at,
            "updated_at": subnet.updated_at,
            **stats
        }
        result.append(subnet_dict)

    return result


@router.get("/subnets/{subnet_id}", response_model=IPSubnetResponse)
async def get_subnet(
    subnet_id: int,
    db: AsyncSession = Depends(get_db)
):
    """Get subnet by ID"""
    subnet = await ipam_service.get_subnet(db, subnet_id)
    if not subnet:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Subnet {subnet_id} not found"
        )

    # Add statistics
    stats = await ipam_service.get_subnet_statistics(db, subnet.id)
    return {
        "id": subnet.id,
        "name": subnet.name,
        "network": str(subnet.network),
        "description": subnet.description,
        "vlan_id": subnet.vlan_id,
        "gateway": str(subnet.gateway) if subnet.gateway else None,
        "dns_servers": subnet.dns_servers,
        "enabled": subnet.enabled,
        "auto_scan": subnet.auto_scan,
        "scan_interval": subnet.scan_interval,
        "last_scan_at": subnet.last_scan_at,
        "created_at": subnet.created_at,
        "updated_at": subnet.updated_at,
        **stats
    }


@router.put("/subnets/{subnet_id}", response_model=IPSubnetResponse)
async def update_subnet(
    subnet_id: int,
    update_data: IPSubnetUpdate,
    db: AsyncSession = Depends(get_db)
):
    """Update an existing subnet"""
    try:
        subnet = await ipam_service.update_subnet(
            db=db,
            subnet_id=subnet_id,
            name=update_data.name,
            description=update_data.description,
            vlan_id=update_data.vlan_id,
            gateway=str(update_data.gateway) if update_data.gateway else None,
            dns_servers=update_data.dns_servers,
            enabled=update_data.enabled,
            auto_scan=update_data.auto_scan,
            scan_interval=update_data.scan_interval
        )

        if not subnet:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Subnet {subnet_id} not found"
            )

        # Add statistics
        stats = await ipam_service.get_subnet_statistics(db, subnet.id)
        return {
            "id": subnet.id,
            "name": subnet.name,
            "network": str(subnet.network),
            "description": subnet.description,
            "vlan_id": subnet.vlan_id,
            "gateway": str(subnet.gateway) if subnet.gateway else None,
            "dns_servers": subnet.dns_servers,
            "enabled": subnet.enabled,
            "auto_scan": subnet.auto_scan,
            "scan_interval": subnet.scan_interval,
            "last_scan_at": subnet.last_scan_at,
            "created_at": subnet.created_at,
            "updated_at": subnet.updated_at,
            **stats
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating subnet: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to update subnet: {str(e)}"
        )


@router.get("/subnets/{subnet_id}/statistics", response_model=IPSubnetStatistics)
async def get_subnet_statistics(
    subnet_id: int,
    db: AsyncSession = Depends(get_db)
):
    """Get subnet statistics"""
    subnet = await ipam_service.get_subnet(db, subnet_id)
    if not subnet:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Subnet {subnet_id} not found"
        )

    stats = await ipam_service.get_subnet_statistics(db, subnet.id)
    return {
        'subnet_id': subnet.id,
        'subnet_name': subnet.name,
        'network': str(subnet.network),
        **stats,
        'last_scan_at': subnet.last_scan_at
    }


@router.delete("/subnets/{subnet_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_subnet(
    subnet_id: int,
    db: AsyncSession = Depends(get_db)
):
    """Delete a subnet and all its IP addresses"""
    subnet = await ipam_service.get_subnet(db, subnet_id)
    if not subnet:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Subnet {subnet_id} not found"
        )

    await db.delete(subnet)
    await db.commit()


# IP Address endpoints
@router.get("/ip-addresses", response_model=IPAddressListResponse)
async def list_ip_addresses(
    subnet_id: Optional[int] = Query(None),
    status: Optional[IPStatus] = Query(None),
    is_reachable: Optional[bool] = Query(None),
    search: Optional[str] = Query(None),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=10000),
    db: AsyncSession = Depends(get_db)
):
    """
    List IP addresses with filters and pagination

    - **subnet_id**: Filter by subnet
    - **status**: Filter by status (available, used, reserved, offline)
    - **is_reachable**: Filter by reachability
    - **search**: Search in IP, hostname, or description
    - **skip**: Number of records to skip (pagination)
    - **limit**: Maximum number of records to return (1-10000)
    """
    from sqlalchemy import select, and_, or_
    from models.switch import Switch

    # Build query with switch join
    query = select(IPAddress, Switch.name.label('switch_name')).outerjoin(
        Switch, IPAddress.switch_id == Switch.id
    )

    # Apply filters
    conditions = []
    if subnet_id:
        conditions.append(IPAddress.subnet_id == subnet_id)
    if status:
        conditions.append(IPAddress.status == status)
    if is_reachable is not None:
        conditions.append(IPAddress.is_reachable == is_reachable)
    if search:
        conditions.append(
            or_(
                cast(IPAddress.ip_address, String).contains(search),
                IPAddress.hostname.ilike(f'%{search}%'),
                IPAddress.dns_name.ilike(f'%{search}%'),
                IPAddress.system_name.ilike(f'%{search}%'),
                IPAddress.machine_type.ilike(f'%{search}%'),
                IPAddress.vendor.ilike(f'%{search}%'),
                cast(IPAddress.mac_address, String).ilike(f'%{search}%'),
                IPAddress.description.ilike(f'%{search}%')
            )
        )

    if conditions:
        query = query.where(and_(*conditions))

    # Get total count before pagination
    count_query = select(func.count(IPAddress.id))
    if conditions:
        count_query = count_query.where(and_(*conditions))

    count_result = await db.execute(count_query)
    total = count_result.scalar() or 0

    # Apply pagination and ordering
    ip_text = func.host(IPAddress.ip_address)
    query = query.order_by(
        cast(func.split_part(ip_text, '.', 1), Integer),
        cast(func.split_part(ip_text, '.', 2), Integer),
        cast(func.split_part(ip_text, '.', 3), Integer),
        cast(func.split_part(ip_text, '.', 4), Integer),
    ).offset(skip).limit(limit)

    result = await db.execute(query)
    rows = result.all()

    overlays = await ipam_service.get_live_location_overlays(
        db,
        [ip_addr for ip_addr, _ in rows]
    )

    # Convert to response format
    items = []
    for ip_addr, switch_name in rows:
        overlay = overlays.get(ip_addr.id, {})
        ip_dict = {
            'id': ip_addr.id,
            'subnet_id': ip_addr.subnet_id,
            'ip_address': str(ip_addr.ip_address),
            'status': ip_addr.status,
            'is_reachable': ip_addr.is_reachable,
            'response_time': ip_addr.response_time,
            'hostname': ip_addr.hostname,
            'hostname_source': ip_addr.hostname_source,
            'dns_name': ip_addr.dns_name,
            'system_name': ip_addr.system_name,
            'mac_address': str(ip_addr.mac_address) if ip_addr.mac_address else overlay.get('mac_address'),
            'vendor': ip_addr.vendor,
            'machine_type': ip_addr.machine_type,
            'switch_id': ip_addr.switch_id or overlay.get('switch_id'),
            'switch_name': switch_name or overlay.get('switch_name'),
            'switch_port': ip_addr.switch_port or overlay.get('switch_port'),
            'vlan_id': ip_addr.vlan_id or overlay.get('vlan_id'),
            'os_type': ip_addr.os_type,
            'os_name': ip_addr.os_name,
            'os_version': ip_addr.os_version,
            'os_vendor': ip_addr.os_vendor,
            'description': ip_addr.description,
            'last_boot_time': ip_addr.last_boot_time,
            'last_seen_at': ip_addr.last_seen_at,
            'last_scan_at': ip_addr.last_scan_at,
            'scan_count': ip_addr.scan_count,
            'created_at': ip_addr.created_at,
            'updated_at': ip_addr.updated_at
        }
        items.append(ip_dict)

    return IPAddressListResponse(items=items, total=total)


@router.get("/ip-addresses/{ip_id}", response_model=IPAddressDetailResponse)
async def get_ip_address(
    ip_id: int,
    db: AsyncSession = Depends(get_db)
):
    """Get IP address details by ID"""
    from sqlalchemy import select
    from models.ipam import IPSubnet
    from models.switch import Switch

    result = await db.execute(
        select(IPAddress, IPSubnet.name, Switch.name)
        .join(IPSubnet, IPAddress.subnet_id == IPSubnet.id)
        .outerjoin(Switch, IPAddress.switch_id == Switch.id)
        .where(IPAddress.id == ip_id)
    )
    row = result.first()

    if not row:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"IP address {ip_id} not found"
        )

    ip_addr, subnet_name, switch_name = row
    overlay = (await ipam_service.get_live_location_overlays(db, [ip_addr])).get(ip_addr.id, {})
    return {
        'id': ip_addr.id,
        'subnet_id': ip_addr.subnet_id,
        'ip_address': str(ip_addr.ip_address),
        'status': ip_addr.status,
        'is_reachable': ip_addr.is_reachable,
        'response_time': ip_addr.response_time,
        'hostname': ip_addr.hostname,
        'hostname_source': ip_addr.hostname_source,
        'dns_name': ip_addr.dns_name,
        'system_name': ip_addr.system_name,
        'machine_type': ip_addr.machine_type,
        'vendor': ip_addr.vendor,
        'contact': ip_addr.contact,
        'location': ip_addr.location,
        'mac_address': str(ip_addr.mac_address) if ip_addr.mac_address else overlay.get('mac_address'),
        'os_type': ip_addr.os_type,
        'os_name': ip_addr.os_name,
        'os_version': ip_addr.os_version,
        'os_vendor': ip_addr.os_vendor,
        'switch_id': ip_addr.switch_id or overlay.get('switch_id'),
        'switch_port': ip_addr.switch_port or overlay.get('switch_port'),
        'vlan_id': ip_addr.vlan_id or overlay.get('vlan_id'),
        'last_seen_at': ip_addr.last_seen_at,
        'last_boot_time': ip_addr.last_boot_time,
        'last_scan_at': ip_addr.last_scan_at,
        'scan_count': ip_addr.scan_count,
        'description': ip_addr.description,
        'created_at': ip_addr.created_at,
        'updated_at': ip_addr.updated_at,
        'subnet_name': subnet_name,
        'switch_name': switch_name or overlay.get('switch_name')
    }


@router.put("/ip-addresses/{ip_id}", response_model=IPAddressResponse)
async def update_ip_address(
    ip_id: int,
    update_data: IPAddressUpdate,
    db: AsyncSession = Depends(get_db)
):
    """Update IP address"""
    ip_addr = await ipam_service.update_ip_address(
        db=db,
        ip_id=ip_id,
        status=update_data.status,
        hostname=update_data.hostname,
        mac_address=update_data.mac_address,
        description=update_data.description
    )

    if not ip_addr:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"IP address {ip_id} not found"
        )

    return ip_addr


@router.get("/ip-addresses/{ip_id}/history", response_model=List[IPScanHistoryResponse])
async def get_ip_scan_history(
    ip_id: int,
    limit: int = Query(50, ge=1, le=500),
    db: AsyncSession = Depends(get_db)
):
    """Get scan history for an IP address"""
    from sqlalchemy import select
    from models.ipam import IPScanHistory

    result = await db.execute(
        select(IPScanHistory)
        .where(IPScanHistory.ip_address_id == ip_id)
        .order_by(IPScanHistory.scanned_at.desc())
        .limit(limit)
    )

    return result.scalars().all()


# Scan endpoints
@router.get("/scan-status", response_model=IPAMScanStatusResponse)
async def get_scan_status():
    """Return the latest IPAM scan status snapshot."""
    return ipam_scan_status_service.get_status()


@router.get("/scan-events")
async def stream_scan_events(request: Request):
    """Stream live IPAM scan progress via Server-Sent Events."""
    async def event_generator():
        queue = await ipam_scan_status_service.subscribe()
        try:
            while True:
                if await request.is_disconnected():
                    break

                try:
                    event = await asyncio.wait_for(queue.get(), timeout=15)
                    yield f"data: {json.dumps(event, default=str, ensure_ascii=False)}\n\n"
                except asyncio.TimeoutError:
                    yield ": keep-alive\n\n"
        finally:
            await ipam_scan_status_service.unsubscribe(queue)

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        }
    )


@router.post("/scan-stream", response_model=IPAMScanStartResponse)
async def start_scan_stream(
    scan_request: IPScanRequest,
    db: AsyncSession = Depends(get_db)
):
    """
    Start an async subnet scan and watch progress through `/scan-events`.

    This is intended for the SolarWinds-style UI flow where the request returns
    immediately and progress is shown live at the bottom of the page.
    """
    if not scan_request.subnet_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="scan-stream 仅支持子网扫描，请提供 subnet_id"
        )

    if network_scheduler.is_ipam_scan_running():
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=network_scheduler.get_ipam_scan_busy_message()
        )

    subnet = await ipam_service.get_subnet(db, scan_request.subnet_id)
    if not subnet:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Subnet {scan_request.subnet_id} not found"
        )

    subnet_label = str(subnet.network)
    await network_scheduler._ipam_scan_lock.acquire()

    try:
        network_scheduler.set_ipam_scan_context(
            f"当前正在手动扫描子网 {subnet_label}"
        )
        session_id = await ipam_scan_status_service.start_scan(
            source="manual",
            scan_type=scan_request.scan_type,
            total_subnets=1,
            message=f"已启动手动扫描，准备扫描子网 {subnet_label}"
        )
    except Exception:
        network_scheduler.clear_ipam_scan_context()
        network_scheduler._ipam_scan_lock.release()
        raise

    async def run_manual_scan() -> None:
        manual_db = AsyncSessionLocal()
        try:
            summary = await ipam_service.scan_subnet(
                db=manual_db,
                subnet_id=scan_request.subnet_id,
                scan_type=scan_request.scan_type,
                include_results=False,
                progress_callback=ipam_scan_status_service.consume_scan_event,
                subnet_index=1,
                total_subnets=1
            )
            await ipam_scan_status_service.complete_scan(
                summary=summary,
                message=(
                    f"子网 {subnet_label} 扫描完成，"
                    f"在线 {summary.get('reachable', 0)}，"
                    f"离线 {summary.get('unreachable', 0)}"
                )
            )
        except Exception as exc:
            logger.error(f"Async subnet scan failed for {subnet_label}: {exc}", exc_info=True)
            await ipam_scan_status_service.fail_scan(
                error=str(exc),
                message=f"子网 {subnet_label} 扫描失败"
            )
        finally:
            await manual_db.close()
            network_scheduler.clear_ipam_scan_context()
            network_scheduler._ipam_scan_lock.release()

    asyncio.create_task(run_manual_scan())

    return {
        "session_id": session_id,
        "message": f"已启动后台扫描：{subnet_label}",
        "status": ipam_scan_status_service.get_status()
    }


@router.post("/scan", response_model=IPScanSummary)
async def scan_ips(
    scan_request: IPScanRequest,
    db: AsyncSession = Depends(get_db)
):
    """
    Scan IP addresses - SIMPLIFIED SYNC MODE

    - **subnet_id**: Scan all IPs in a subnet (runs synchronously)
    - **ip_addresses**: Scan specific IP addresses
    - **scan_type**: 'quick' (ping only) or 'full' (ping + hostname + MAC + OS)

    Note: For subnets, this runs synchronously and may take 2-3 minutes.
    Frontend should poll for updated last_scan_at to detect completion.
    """
    if scan_request.subnet_id:
        if network_scheduler.is_ipam_scan_running():
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=network_scheduler.get_ipam_scan_busy_message()
            )

        # Run scan synchronously (reliable but slow)
        async with network_scheduler._ipam_scan_lock:
            network_scheduler.set_ipam_scan_context(
                f"当前正在手动扫描子网 {scan_request.subnet_id}"
            )
            try:
                subnet = await ipam_service.get_subnet(db, scan_request.subnet_id)
                subnet_label = str(subnet.network) if subnet else str(scan_request.subnet_id)
                await ipam_scan_status_service.start_scan(
                    source="manual",
                    scan_type=scan_request.scan_type,
                    total_subnets=1,
                    message=f"已启动同步扫描，准备扫描子网 {subnet_label}"
                )
                result = await ipam_service.scan_subnet(
                    db=db,
                    subnet_id=scan_request.subnet_id,
                    scan_type=scan_request.scan_type,
                    progress_callback=ipam_scan_status_service.consume_scan_event,
                    subnet_index=1,
                    total_subnets=1
                )
                await ipam_scan_status_service.complete_scan(
                    summary=result,
                    message=(
                        f"子网 {subnet_label} 扫描完成，"
                        f"在线 {result.get('reachable', 0)}，"
                        f"离线 {result.get('unreachable', 0)}"
                    )
                )
            except Exception as exc:
                await ipam_scan_status_service.fail_scan(
                    error=str(exc),
                    message=f"子网 {scan_request.subnet_id} 扫描失败"
                )
                raise
            finally:
                network_scheduler.clear_ipam_scan_context()
        return result

    elif scan_request.ip_addresses:
        # Scan specific IPs
        from services.ip_scan import ip_scan_service
        import time

        start_time = time.time()
        results = await ip_scan_service.scan_multiple_ips(
            scan_request.ip_addresses,
            scan_request.scan_type
        )
        elapsed = time.time() - start_time

        reachable = sum(1 for r in results if r['is_reachable'])

        return {
            'total_scanned': len(results),
            'reachable': reachable,
            'unreachable': len(results) - reachable,
            'new_devices': 0,
            'changed_devices': 0,
            'scan_duration': elapsed,
            'results': results
        }
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Either subnet_id or ip_addresses must be provided"
        )


# Dashboard endpoint
@router.get("/dashboard", response_model=IPAMDashboard)
async def get_dashboard(
    db: AsyncSession = Depends(get_db)
):
    """Get IPAM dashboard statistics"""
    return await ipam_service.get_dashboard_stats(db)


# Network Search endpoint
@router.post("/network-search", response_model=NetworkSearchResponse)
async def search_network(
    request: NetworkSearchRequest,
    db: AsyncSession = Depends(get_db)
):
    """
    Search for all IPs in a given network (CIDR)

    Example: 10.101.63.0/24

    Returns all IP addresses in the subnet and their usage status.
    If the network exists in IPAM, returns managed IP data.
    If not, returns calculated network information.
    """
    try:
        result = await ipam_service.search_network(db, request.network)
        return result
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        logger.error(f"Error searching network: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to search network"
        )


# Subnet Calculator endpoint
@router.post("/subnet-calculator", response_model=SubnetCalculatorResponse)
async def calculate_subnet(
    request: SubnetCalculatorRequest
):
    """
    Calculate subnet information from IP address and netmask

    Input:
    - IP address (e.g., 10.101.63.25)
    - Netmask (e.g., 255.255.255.0 or /24)

    Output:
    - Network address
    - Broadcast address
    - First/last usable host
    - Total hosts
    - Subnet mask
    - CIDR notation
    - Binary and hex representations
    """
    try:
        from services.ipam_service import calculate_subnet_info
        result = calculate_subnet_info(request.ip_address, request.netmask)
        return result
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        logger.error(f"Error calculating subnet: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to calculate subnet"
        )
